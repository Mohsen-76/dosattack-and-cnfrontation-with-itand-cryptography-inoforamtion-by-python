"""
written by :

 _      ____  _     ____  _____ _       
/ \__/|/  _ \/ \ /|/ ___\/  __// \  /|  
| |\/||| / \|| |_|||    \|  \  | |\ ||  
| |  ||| \_/|| | ||\___ ||  /_ | | \||  
\_/  \|\____/\_/ \|\____/\____\\_/  \|  
                                        
name of project : dosattack-and-cnfrontation-with-itand-cryptography-inoforamtion-by-python
date of create : 12 March 2022

"""


import socket
import os
import time
import sys
import threading
import subprocess
 
import random

from datetime import datetime

try:
    import mysql.connector
except:
    os.system('pip install mysql-connector-python')


try:
    from cryptography.fernet import Fernet
except:
    os.system('pip install cryptography')   

try:
    from openpyxl import Workbook,load_workbook
    from openpyxl.utils import get_column_letter
except:
    os.system('pip install openpyxl')    

try:
    from colorama import Fore,init
    init()
except:
    os.system('pip install colorama')    

try:
    import networkx as nx
except:
    os.system('pip install networkx')        




# variable
ip=''
port=0
server=None
client=None
addr=''
database=''

all_connection=[]
# menu="""
# [b] BROADCASRT
# [u] UNICAST
# [g] GRAPH
# """

logo="""

███████╗███████╗██████╗ ██╗   ██╗███████╗██████╗ 
██╔════╝██╔════╝██╔══██╗██║   ██║██╔════╝██╔══██╗
███████╗█████╗  ██████╔╝██║   ██║█████╗  ██████╔╝
╚════██║██╔══╝  ██╔══██╗╚██╗ ██╔╝██╔══╝  ██╔══██╗
███████║███████╗██║  ██║ ╚████╔╝ ███████╗██║  ██║
╚══════╝╚══════╝╚═╝  ╚═╝  ╚═══╝  ╚══════╝╚═╝  ╚═╝
                                                 
"""
print(Fore.RED+logo)
time.sleep(0.1)
print('-----------------------------------------------')
time.sleep(0.1)


# database


mydb=mysql.connector.connect(
    host='localhost',
    user='admin',
    password='majic',
    database='network'
)
mycursor=mydb.cursor()

while True:
    octed_status=False
    try:
        print(Fore.YELLOW+'')
        ip = input("┌─["+"ENTER IP OF SERVER"+"""]
└──╼ """+"卐 ")
        if ip==None or ip=="" or ip=="\n":
            print(Fore.RED+'the ip is empty'.upper())
            continue
        # 192.168.1.200
        condition=str(ip).split('.')
        for dot in condition:
            if not dot.isdigit():
                octed_status=True
                break
        if octed_status:
            print(Fore.RED+'one of the octed is not correct'.upper())
            continue


        if len(condition)!=4:
            print(Fore.RED+'your ip is not correct '.upper())
            continue

        break

    except KeyboardInterrupt:
        sys.exit()    
    except:
        print(Fore.RED+'i cant get the ip'.upper())    


# print(ip)
time.sleep(0.1)
while True:
    try:
        print(Fore.YELLOW+'')
        port = input("┌─["+"ENTER PORT OF SERVER"+"""]
└──╼ """+"卐 ")
       
        if not port.isdigit():
            print(Fore.RED+'number of port is not number'.upper())
            continue
        port=int(port)

        if port<1000:
            print(Fore.RED+'number of port has to be more 1000'.upper())
            continue

        
        break

    except KeyboardInterrupt:
        sys.exit()    
    except:
        print(Fore.RED+'i cant get the port'.upper())  

print('-----------------------------------------------')
time.sleep(0.1)

 



def check_with_ping(ip):
    try:
        res=subprocess.check_output(f'ping {ip}',shell=True,text=True,stderr=subprocess.STDOUT)
        if 'Sent = 4, Received = 4, Lost = 0' in res and (int(ip.split('.')[0])>=192 ) :
            return True
        else:
            return False  
    except:
            return False       

def read_query(query):
    try:
        mycursor.execute(query)
        result=mycursor.fetchall()
        return result
    except Exception as er:
        print(er) 

def execute_to_database(query):
    try:
        mycursor.execute(query)
        mydb.commit()
    except Exception as er:
        print(er)



query_read_ip='''select ip from info'''


query_del="delete from info where id=1"  
query_create_database='create database network'
query_table="""
create table info(
    ip varchar(255) ,
    gmail varchar(255) ,
    auth_code int ,
    hash_code varchar(255),
    is_block int,
    is_login int
)
""" 
query_table_log="""
create table logs(
    source varchar(255),
    destination varchar(255),
    data varchar(255),
    time varchar(255)
)
"""
query_insert="""insert into info (
    ip,
    gmail,
    auth_code,
    hash_code ,
    is_login,
    is_block
)values("{}","{}",{},"{}",{},{})
"""
query_insert_log="""insert into logs (
    source ,
    destination,
    data,
    time
)values("{}","{}","{}","{}")
"""
query_read_hash_code="""select hash_code from info where ip='{}'"""
query_read_auth_code='''select auth_code from info where ip="{}"'''
query_read_is_login='''select is_login from info where ip="{}"'''
query_read_is_block='''select is_block from info where ip="{}"'''

query_update_is_login="""update info set is_login={} where ip='{}'"""
query_update_is_block="""update info set is_block={} where ip='{}'"""
# create database
execute_to_database(query_create_database)

# create table info
execute_to_database(query_table)
# create table logs
execute_to_database(query_table_log)

def create_auth_and_hash_code():
    auth=random.randint(1000,9999)
    hash=Fernet.generate_key()
    return auth,hash

def send_email(user, pwd, recipient, subject, body):
    import smtplib

    FROM = user
    TO = recipient if isinstance(recipient, list) else [recipient]
    SUBJECT = subject
    TEXT = body

    # Prepare actual message
    message = """From: %s\nTo: %s\nSubject: %s\n\n%s
    """ % (FROM, ", ".join(TO), SUBJECT, TEXT)
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.ehlo()
        server.starttls()
        server.login(user, pwd)
        server.sendmail(FROM, TO, message)
        server.close()
        print ('successfully sent the mail')
    except:
        print ("failed to send mail")

def give_ip(query):
    ip_list=[]
    ip_ok=read_query(query)
    for ip_tup in ip_ok:
        for ip_ in tuple(ip_tup):
            if len(str(ip_).split('.'))==4:
                ip_list.append(ip_)
    return ip_list            

def give_is_login(query):
    try: 
        is_login_code=read_query(query)
        # print(is_login_code[0][0])
        return int (is_login_code[0][0])
    except:
        return 0    

def give_is_block(query):
    try:
        is_block_code=read_query(query)
        # print(is_block_code[0][0])
        return int (is_block_code[0][0])
    except:
        return 0

def give_auth_code(query):
    try:
        auth_codes=read_query(query)
        # print(auth_codes[0][0])
        return int(auth_codes[0][0])
    except:
        print('not found auth code')    

def give_hash_code(query):
    try:
        hash_codes=read_query(query)
        # print(hash_codes[0][0])
        return hash_codes[0][0]
    except:
        print('not found hash code')    

def update(query):
    try:
        mycursor.execute(query)
        mydb.commit()
    except Exception as er:
        print(er)



ip_dic={}
is_vorod=False
def entry_user(client,ip):
    global is_vorod,ip_dic
    # print(is_vorod)
    global query_insert,query_read_auth_code,query_update_is_login,query_read_is_login,query_read_is_block
    gmail_info=open('userandpass.txt','r').read().split('\n')
    username,password=gmail_info[0],gmail_info[1]

    if not check_with_ping(str(ip)):
        print('ip is not in my network ...')
        return
   
    ip_list=give_ip(query_read_ip)
    print(ip_list)

    read_login=query_read_is_login.format(str(ip))
    is_login_ok=give_is_login(read_login)

    
    if  is_login_ok==0:
        ip_dic[str(ip)].send('gmail'.encode())
        gmail=ip_dic[str(ip)].recv(1234).decode()
        if gmail=='again':
            pass
        else:
            while True:
                auth_code,hash_code=create_auth_and_hash_code()
                body=f"auth code : {auth_code} \n hash code : {hash_code}"
                send_email(username,password,gmail,'authentication',body)
                is_reciave=ip_dic[str(ip)].recv(1234).decode()
                if is_reciave=='not reciave':
                    continue
                else:
                    break

            a_insert=query_insert.format(ip,gmail,auth_code,hash_code.decode(),0,0)
            execute_to_database(a_insert)
            a_code=query_read_auth_code.format(ip)
            auth_code_in_databse=give_auth_code(a_code)
            while True:
                code=ip_dic[str(ip)].recv(1234).decode()
                if int(code)==auth_code_in_databse:
                    ip_dic[str(ip)].send('correct'.encode())
                    u_login=query_update_is_login.format(1,ip)
                    update(u_login)

                    r_login=query_read_is_login.format(str(ip))
                    is_login_ok=give_is_login(r_login)
                    if is_login_ok==1:
                        ip_dic[str(ip)].send('welcome'.encode())
                        # print('welcome')
                        # if len(ip_dic)!=2:
                        #     print('two client have to connect to me')
                        #     return
                        if not is_vorod:
                            is_vorod=True
                            panel_server()
                            
                    break
                else:
                    ip_dic[str(ip)].send('wrong'.encode())   
                    continue
    elif is_login_ok==1:
        # r_login=query_read_is_login.format(str(ip))
        # is_login_ok=give_is_login(r_login)
        # if is_login_ok==1:
            ip_dic[str(ip)].send('welcome'.encode())
            print('welcome')
            # if len(ip_dic)!=2:
            #     print('two client have to connect to me')
            #     return
            if not is_vorod:
                is_vorod=True
                panel_server()
            

menu="""
[1] broadcast
[2] unicast
[3] send between clients
[4] graph
"""


def show_graph(ip_dic):
    import matplotlib.pyplot as plt
    g=nx.Graph()
    g.add_node('server')
    for ip in ip_dic.keys():
        g.add_node(ip)
        g.add_edge(ip,'server')    
    nx.draw(g,with_labels=True, font_weight='bold',node_size =[8000])
    plt.show()

def broadcast_send(ip_dic):
    # ip_dic[ip].send('broad'.encode())
    for client in ip_dic.values():
        client.send('broad'.encode())
    data=show_xlsx_file()
    for ip,cli in ip_dic.items():
        insert_log=query_insert_log.format('server',ip,data,str(datetime.now()))
        execute_to_database(insert_log)
        hash_query=query_read_hash_code.format(str(ip))
        hash_code=give_hash_code(hash_query).encode()
        # print(type(hash_code))
        f=Fernet(hash_code)

        data_encrypt=f.encrypt(data.encode())
        # print(data_encrypt)
        cli.send(data_encrypt)
    for ip,cli in ip_dic.items():
        message=cli.recv(12345).decode()
        if message=='true':
            print(f'{ip} says seccusful to recaive data'.upper()) 
        elif message=='false':
            print(f'{ip} says fail to recaive data'.upper()) 

def unicast_send(ip_dic):
    # data=show_xlsx_file()
    host=input(Fore.YELLOW+'enter ip of client : '.upper())
    if host in ip_dic.keys():   
        data=show_xlsx_file()
        ip_dic[host].send('uni'.encode())
        insert_log=query_insert_log.format('server',host,data,str(datetime.now()))
        execute_to_database(insert_log)
        hash_query=query_read_hash_code.format(str(host))
        hash_code=give_hash_code(hash_query).encode()
        # print(type(hash_code))
        f=Fernet(hash_code)
        data_encrypt=f.encrypt(data.encode())
        # print(data_encrypt)
        ip_dic[host].send(data_encrypt)
        message=ip_dic[host].recv(12345).decode()
        if message=='true':
                print(f'{host} says seccusful to recaive data'.upper()) 
        elif message=='false':
                print(f'{host} says fail to recaive data'.upper())
    else:
        print(Fore.RED+'there is not client with ip : '.upper(),host)  

def show_xlsx_file():
  
    import os
    # show all xlsx file ... 
    xlsx_list=[]
    all_file=os.listdir(os.getcwd())
    for file in all_file:
        sp =file.split('.')
        if len(sp)>1:
            if sp[1]=='xlsx':
                xlsx_list.append(file)

    while True:    
        for index , file_xlsx in enumerate(xlsx_list,start=1):
            print(Fore.YELLOW+f'[ {index} ] {file_xlsx}') 
        i=input(Fore.GREEN+'choose a number >>>> '.upper())
        if i==None or i=='' or i==' ' or i=='\n':
            print(Fore.RED+'empty data !!!'.upper())
            show_xlsx_file()
        if not i.isdigit():
            print(Fore.RED+'not number !!!'.upper())
            show_xlsx_file()
        i=int(i)
        break
    print(Fore.BLUE+f'{xlsx_list[i-1]} choosed ... ')  
    # load data from xlsx file ...
    wb=load_workbook(xlsx_list[i-1]) 
    data=''
    for _index,sheet in enumerate( wb.sheetnames):
        var=''
        ws=wb[sheet]
        for row in range(1,ws.max_row+1):
            for col in range(1,ws.max_column+1):
                char =get_column_letter(col)
                value=ws[char+str(row)].value
                if col!=ws.max_column:
                    var+=str(value)+','
                else:
                    var+=str(value)+'\n' 
        if _index!=len(wb.sheetnames)-1:                 
            data=data+var+'-----'+'\n'
        else:
            data=data+var    

    wb.save(xlsx_list[i-1]) 
    return data

def give_ip_between(mess):
    while True:
        octed_status=False
        ip = input(f'enter the ip of {mess} : ')
        if ip==None or ip=="" or ip=="\n":
            continue
        condition=str(ip).split('.')
        for dot in condition:
            if not dot.isdigit():
                octed_status=True
                break
        if octed_status:
            continue

        if len(condition)!=4:
            continue

        break
    return ip

def send_between_clients(ip_dic):
    if len(ip_dic)>=2:
        src=give_ip_between('source')
        dst=give_ip_between('destination')
        if src in ip_dic.keys() and dst in ip_dic.keys():
            ip_dic[src].send('send'.encode())
            ip_dic[dst].send('recv'.encode()) 
            # give key src from database
            hash_query_src=query_read_hash_code.format(str(src))
            hash_for_src=give_hash_code(hash_query_src).encode()
            f_src=Fernet(hash_for_src)
            # give data from src
            data_from_src=f_src.decrypt(ip_dic[src].recv(12345678)).decode()
            
            # set to data base
            insert_log=query_insert_log.format(src,dst,data_from_src,str(datetime.now()))
            execute_to_database(insert_log)
            # give key dst from data base
            hash_query_dst=query_read_hash_code.format(str(dst))
            hash_for_dst=give_hash_code(hash_query_dst).encode()
            f_dst=Fernet(hash_for_dst)
            # send data to dst
            data_dst=f_dst.encrypt(data_from_src.encode())
            ip_dic[dst].send(data_dst)

        else:
            print('there arnt that ips in my network ...')    

    else:
        print('more one client have to be online ...')

# server panel
def panel_server():
 
    try:
        while True:
            global ip_dic
            print(ip_dic)
            print(menu)
            choose_num=input('choose a number : ')
            if choose_num=='' or choose_num==None or choose_num=='\n':
                continue
            if not choose_num.isdigit():
                continue
            choose_num=int(choose_num)
       
            # broadcast
            if choose_num==1:
                broadcast_send(ip_dic)
            # unicast
            elif choose_num==2:
                unicast_send(ip_dic)
            # send between clients
            elif choose_num==3:
                send_between_clients(ip_dic)
            # graph
            elif choose_num==4:
                show_graph(ip_dic)

        
    except:
        pass    

# regular attack
# get mines between each item and if all mineses are the same it is regular attack 
def get_time_regular_attack(list):
    print('i am in get time regualr function')
    print(list)
    help_list=[]
    for i in range(len(list)):
        if i !=len(list)-1:
            if abs(list[i+1])>abs(list[i]):
                 help_list.append(abs(list[i+1])-abs(list[i]))
    st_list=[]
    for j in range(len(help_list)):
        if j!=len(help_list)-1:
            if help_list[j]==help_list[j+1]:
                st_list.append(True)
            else:
                st_list.append(False)    
        else:
            if help_list[j]==help_list[j-1]:
                st_list.append(True)
            else:
                st_list.append(False)          

    if all(st_list) :
        print(st_list)
        return True
    else:
        print(st_list)
        return False  

# get seconds in each 5 requset from every ip 
def get_time(time1,time2):
    time1=str(time1).split(':')
    time2=str(time2).split(':')
    if len(time1)==3 and len(time2)==3:
        min1=int(time1[1])
        sec1=int(time1[2])

        min2=int(time2[1])
        sec2=int(time2[2])
        if min1==min2:
            return  sec2-sec1
        else:
            seconds=0
            for m  in range(min1,min2):
                for s in range(sec1,60):
                    seconds+=1
                sec1=0
            seconds+=sec2
            return seconds

# setecting dos attack
dic_ip={}
def detect_dos_attack(client_host):
    ttm=str(datetime.now().hour)+':'+str(datetime.now().minute)+':'+str(datetime.now().second)
    
    if str(client_host) not in dic_ip.keys():
        dic_ip[str(client_host)]={'how_many':0,'start':'','end':'','st_block':False,'each_time':[]}
    else:

        if dic_ip[str(client_host)]['st_block']:
            updata_block=query_update_is_block.format(1,str(client_host))
            execute_to_database(updata_block)
        
        dic_ip[str(client_host)]['how_many']+=1   
        print(dic_ip)
        if len(dic_ip[str(client_host)]['each_time'])<=10:
             print('adding')
             dic_ip[str(client_host)]['each_time'].append(int(str(ttm).split(':')[2]))
        else:
            print('checking')
            st=get_time_regular_attack(dic_ip[str(client_host)]['each_time'])
            print(st)
            if dic_ip[str(client_host)]['how_many']>10 and st :
                print('regular attack handle')
                dic_ip[str(client_host)]['st_block']=True



        if dic_ip[str(client_host)]['how_many']==1:
            dic_ip[str(client_host)]['start']=ttm
        elif dic_ip[str(client_host)]['how_many']%5==0:
            dic_ip[str(client_host)]['end']=ttm
            # 12:30:12      ,      12:30:28
            get_sec=get_time(dic_ip[str(client_host)]['start'],dic_ip[str(client_host)]['end'])
            if get_sec<10:
                dic_ip[str(client_host)]['st_block']=True
            dic_ip[str(client_host)]['start']=ttm   


# create and run honey pot for drop hacker to fake server 
honey_soket=None
def run_server_honey_pot():
    global ip,honey_soket
    try:
        honey_soket=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
        honey_soket.bind((ip,9000))    
        honey_soket.listen(5)
        print(f'honey pot on running {ip}  and port : 9000')
        
    except Exception as er:
        print(er)
        print('some thing is wrong in honey pot')        

# waiting for connect eny client to server but the port changed and hacker drop in sever and cant do any thing
def honey_take(ip):
    global honey_soket
    print(f'{ip} in honey pot .....')
    while True:
        try:
             cli_honey,addr_honney=honey_soket.accept()
             print(f' ip {addr_honney[0]} in dom .....')
        except Exception as er:
            print(er)
   
# run main server on locak ip and port 8000
def run_server_main():
    global ip_dic
    try:
        server=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
        server.bind((ip,int(port)))
        server.listen(4)
        print(Fore.GREEN+'server with ip  : ',server.getsockname()[0],' is running on port : '.upper(),port)
        
    except KeyboardInterrupt:
        sys.exit()    
    except:
        print(Fore.RED+'i cant bind the server'.upper())  

    while True:
        try:
            client,addr=server.accept()
            detect_dos_attack(addr[0])

            if addr[0] not in ip_dic.keys():
                ip_dic[addr[0]]=client

            read_block=query_read_is_block.format(str(addr[0]))
            is_block_ok=give_is_block(read_block)
            if is_block_ok==1: 
                print(f'{addr[0]} blocked ........' )
                if client in ip_dic.values():
                    client.send('block'.encode())
                    # honey_take()
                else :
                    honey_take(addr[0])   
            else:
                threading._start_new_thread(entry_user,(client,addr[0])) 

        except:
            print('no client cant connect to me')    

# run both server in the same time
# threading for main server and run 
tmain=threading.Thread(target=run_server_main)
tmain.start()
# threading for honeypot server and run it 
thoney=threading.Thread(target=run_server_honey_pot)
thoney.start()




