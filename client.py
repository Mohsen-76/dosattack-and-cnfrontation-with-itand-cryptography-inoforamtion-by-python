"""
written by :

 _      ____  _     ____  _____ _       
/ \__/|/  _ \/ \ /|/ ___\/  __// \  /|  
| |\/||| / \|| |_|||    \|  \  | |\ ||  
| |  ||| \_/|| | ||\___ ||  /_ | | \||  
\_/  \|\____/\_/ \|\____/\____\\_/  \|  
                                        
name of project : dosattack-and-cnfrontation-with-itand-cryptography-inoforamtion-by-python
date of create : 12 March 2022
"""


from colorama import Fore,init
import socket
import os
import time
import sys
import threading
import random
import re
try:
    from cryptography.fernet import Fernet
except:
    os.system('pip install cryptography')   

try:
    from colorama import Fore,init
    init()
except:
    os.system('pip isntall colorama')


try:
    from openpyxl import load_workbook,Workbook
    from openpyxl.utils import get_column_letter
except:
    os.system('pip instal openpyxl')    


# variable
ip=''
port=0
server=None
connection=None
chance=0

logo="""

 ██████╗██╗     ██╗███████╗███╗   ██╗████████╗
██╔════╝██║     ██║██╔════╝████╗  ██║╚══██╔══╝
██║     ██║     ██║█████╗  ██╔██╗ ██║   ██║   
██║     ██║     ██║██╔══╝  ██║╚██╗██║   ██║   
╚██████╗███████╗██║███████╗██║ ╚████║   ██║   
 ╚═════╝╚══════╝╚═╝╚══════╝╚═╝  ╚═══╝   ╚═╝   
                                              
"""
print(Fore.RED+logo)
time.sleep(0.1)
print(Fore.CYAN+'-----------------------------------------------')
time.sleep(0.1)
while True:
    octed_status=False
    try:
        print(Fore.YELLOW+'')
        ip = input("┌─["+"ENTER IP OF SERVER"+"""]
└──╼ """+"卐 ")
        if ip==None or ip=="" or ip=="\n":
            print(Fore.RED+'the ip is empty'.upper())
            continue
        condition=str(ip).split('.')
        for dot in condition:
            if not dot.isdigit():
                octed_status=True
                break
        if octed_status:
            print(Fore.RED+'one of the octed is not correct'.upper())
            continue


        if len(condition)!=4:
            print(Fore.RED+'your ip is not correct '.upper())
            continue

        break

    except KeyboardInterrupt:
        sys.exit()    
    except:
        print(Fore.RED+'i cant get the ip'.upper())    


# print(ip)
time.sleep(0.1)
while True:
    try:
        print(Fore.YELLOW+'')
        port = input("┌─["+"ENTER PORT OF SERVER"+"""]
└──╼ """+"卐 ")
        if port==None or port=="" or port=="\n" or port=='0':
            print(Fore.RED+'the port is empty'.upper())
            continue
        if not port.isdigit():
            print(Fore.RED+'number of port is not number'.upper())
            continue
        port=int(port)
        if port<1000:
            print(Fore.RED+'number of port has to be more 1000 '.upper())
            continue
        
        break

    except KeyboardInterrupt:
        sys.exit()    
    except:
        print(Fore.RED+'i cant get the ip'.upper())  

try:
    connection=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
    connection.connect((ip,port))
    print('clinet with ip ',str(connection.getsockname()[0]),' connected to server'.upper())
except:
    print('i can not connect to server '.upper())

def timer():
    for i in range(3,0,-1):
        time.sleep(1)
        print(str(i),end='\r')

def get_key():
    file=open('hash_code.txt','r').read().split('\n')[0]
    key=b''
    if file[0]=='b' and file[1]=="'" and file[-1]=="'":
        key=file[2:-1].encode()
        # print(key)
        # print(type(key))
        return key
    else:
        key=file.encode()
        return key    

def dos_attack():
    global ip,port
    import socket
    import time
    counter=0
    while True:
        time.sleep(1)
        counter+=1
        try:
            s=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
            s.connect((ip,int(port)))
            print(f'i connected {counter}')
        except:
            print('i cant connect')    

def show_xlsx_file():
    # print(' i am in show_xlsx_file')
    import os
    # show all xlsx file ... 
    xlsx_list=[]
    all_file=os.listdir(os.getcwd())
    # print(all_file)
    for file in all_file:
        sp =file.split('.')
        if len(sp)>1:
            if sp[1]=='xlsx':
                xlsx_list.append(file)

    # print(xlsx_list) 
    while True:    
        for index , file_xlsx in enumerate(xlsx_list,start=1):
            print(Fore.YELLOW+f'[ {index} ] {file_xlsx}') 
        i=input(Fore.GREEN+'choose a number >>>> '.upper())
        if i==None or i=='' or i==' ' or i=='\n':
            print(Fore.RED+'empty data !!!'.upper())
            show_xlsx_file()
        if not i.isdigit():
            print(Fore.RED+'not number !!!'.upper())
            show_xlsx_file()
        i=int(i)
        break
    print(Fore.BLUE+f'{xlsx_list[i-1]} choosed ... ')  
    # load data from xlsx file ...
    wb=load_workbook(xlsx_list[i-1]) 
    data=''
    for _index,sheet in enumerate( wb.sheetnames):
        var=''
        ws=wb[sheet]
        for row in range(1,ws.max_row+1):
            for col in range(1,ws.max_column+1):
                char =get_column_letter(col)
                value=ws[char+str(row)].value
                if col!=ws.max_column:
                    var+=str(value)+','
                else:
                    var+=str(value)+'\n' 
        if _index!=len(wb.sheetnames)-1:                 
            data=data+var+'-----'+'\n'
        else:
            data=data+var    

    wb.save(xlsx_list[i-1]) 
    return data


def convert_exel(data):
        sp=data.split('-----') 
        # print(len(sp))
        wb=Workbook()
        # remove all default sheet
        for sheet in wb.sheetnames:
            ws=wb[sheet]
            wb.remove(ws)
        # set name for sheet
        for i in range(len(sp)):
            wb.create_sheet('sheet'+str(i+1))
        # get correct data 
        dic_ex={}
        for index,item in enumerate(sp):
            list_=[]
            xx=re.findall(r'[0-9,]*[0-9]*',item)
            for x in xx:
                if len(x)!=0:
                    list_.append(x)
            dic_ex['sheet'+str(index+1)]=list_
        # print(dic_ex)   

        # set data to exel
        for key,value in dic_ex.items():
            ws=wb[key]
            for val in value:
                val=str(val).split(',')
                ws.append(val)
        wb.save(f'{connection.getsockname()[0]}.xlsx')

while True:
    data=connection.recv(1234).decode()
    if data=='block':
        print('you blocked')
        break

    if data=='gmail':
        while True:
            gmail_input=input('enter gmail : ')
            ack=input('are you sure to send email ? y/n : ')
            if ack=='y':
                connection.send(gmail_input.encode())
                break
            elif ack=='n':
                continue
        print('check your email ....')
        timer()
        while True:
            reciven=input('did you reciave code ? y/n : ')
            if reciven=="y":
                connection.send('reciave'.encode())
                break
            elif reciven=="n":
                connection.send('not reciave'.encode())
                continue
        while True:
            code=input('enter you authentication code : ')
            ack=input('are you sure to send code ? y/n : ')
            if ack=='y':
                connection.send(code.encode())
                status=connection.recv(1234).decode()
                if status=='correct':
                    print('correct')
                    break
                elif status=='wrong':
                    print('wrong')
                    continue
            elif ack=='n':
                continue
    if data=='welcome':
        print('you sign uped')  
    if data=='broad':

        try:
            data_=connection.recv(12345678)
            # print(data_) 
            f=Fernet(get_key())
            data_decode=f.decrypt(data_)
            data_=data_decode.decode()
            convert_exel(data_)
            connection.send('true'.encode())
        except:
            connection.send('false'.encode())   
    if data=='uni':
        try:
            data_=connection.recv(12345678)
            # print(data_) 
            f=Fernet(get_key())
            data_decode=f.decrypt(data_)
            data_=data_decode.decode()
            convert_exel(data_)
            connection.send('true'.encode())
        except:
            connection.send('false'.encode()) 

    if data=='send':
        data_ex=show_xlsx_file()
            
        f=Fernet(get_key())
        mess_encrypt=f.encrypt(data_ex.encode())
        connection.send(mess_encrypt)
        print('message sent ...')

    if data=='recv':
        message_recv=connection.recv(12345678)       
        f=Fernet(get_key())
        mess_decrypt=f.decrypt(message_recv)
        data__=mess_decrypt.decode()
        convert_exel(data__)

            

