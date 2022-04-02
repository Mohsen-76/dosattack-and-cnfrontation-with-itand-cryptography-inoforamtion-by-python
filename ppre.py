import os
try:
    from openpyxl import Workbook,load_workbook
    from openpyxl.utils import get_column_letter
except:
    os.system('pip install openpyxl')    


try:
    from colorama import Fore,init
    init()
except:
    pass

def show_xlsx_file():
    print(' i am in show_xlsx_file')
    import os
    # show all xlsx file ... 
    xlsx_list=[]
    all_file=os.listdir(os.getcwd())
    print(all_file)
    for file in all_file:
        sp =file.split('.')
        if len(sp)>1:
            if sp[1]=='xlsx':
                xlsx_list.append(file)

    print(xlsx_list) 
    while True:    
        for index , file_xlsx in enumerate(xlsx_list,start=1):
            print(Fore.YELLOW+f'[ {index} ] {file_xlsx}') 
        i=input(Fore.GREEN+'choose a number >>>> '.upper())
        if i==None or i=='' or i==' ' or i=='\n':
            print(Fore.RED+'empty data !!!'.upper())
            show_xlsx_file()
        if not i.isdigit():
            print(Fore.RED+'not number !!!'.upper())
            show_xlsx_file()
        i=int(i)
        break
    print(Fore.BLUE+f'{xlsx_list[i-1]} choosed ... ')  
    # load data from xlsx file ...
    wb=load_workbook(xlsx_list[i-1]) 
    data=''
    for _index,sheet in enumerate( wb.sheetnames):
        var=''
        ws=wb[sheet]
        for row in range(1,ws.max_row+1):
            for col in range(1,ws.max_column+1):
                char =get_column_letter(col)
                value=ws[char+str(row)].value
                if col!=ws.max_column:
                    var+=str(value)+','
                else:
                    var+=str(value)+'\n' 
        if _index!=len(wb.sheetnames)-1:                 
            data=data+var+'-----'+'\n'
        else:
            data=data+var    

    wb.save(xlsx_list[i-1]) 
    return data

# import subprocess 
# res =subprocess.check_output('ping 192.168.1.9',shell=True,text=True)
# print(res)